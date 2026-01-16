#!/usr/bin/env python3
"""
AI Samarth CSV Processor with Completion Tracking
Validates files, adds completion columns, and generates summary Excel report
Works on Windows, Linux, and macOS
"""

import csv
import re
import os
from pathlib import Path
from typing import List, Tuple, Dict, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def parse_start_date(date_str: str) -> Optional[datetime.date]:
    """
    Parse Start Date string and return datetime.date object.
    Handles formats: DD/MM/YY, YYYY/MM/DD, and "Not Started"
    
    Returns:
        datetime.date object or None if invalid/not started
    """
    if not date_str or not date_str.strip():
        return None
    
    date_str = date_str.strip()
    
    # Handle "Not Started" case
    if date_str.lower() in ['not started', 'not started', '']:
        return None
    
    # Try DD/MM/YY format (e.g., "23/10/25" or "06/11/25")
    try:
        parts = date_str.split('/')
        if len(parts) == 3:
            # Check if first part is > 31 (likely YYYY/MM/DD format)
            if len(parts[0]) == 4:
                # YYYY/MM/DD format
                year = int(parts[0])
                month = int(parts[1])
                day = int(parts[2])
            else:
                # DD/MM/YY format
                day = int(parts[0])
                month = int(parts[1])
                year = int(parts[2])
                # Convert 2-digit year to 4-digit (assuming 20xx for years < 50, 19xx otherwise)
                if year < 50:
                    year += 2000
                else:
                    year += 1900
            
            # Validate month
            if 1 <= month <= 12 and 1 <= day <= 31:
                return datetime(year, month, day).date()
    except (ValueError, IndexError):
        pass
    
    # Try parsing with datetime
    for fmt in ['%d/%m/%y', '%Y/%m/%d', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
        try:
            dt = datetime.strptime(date_str, fmt).date()
            return dt
        except ValueError:
            continue
    
    return None


def get_month_name(month: int) -> str:
    """Get month name from month number"""
    months = ['', 'January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']
    return months[month] if 1 <= month <= 12 else 'Unknown'


class AISmarthProcessor:
    """Processor for AI Samarth CSV files with completion tracking"""

    # Hard Rules
    COL_R = 17  # Start column (0-indexed) - Pillar
    COL_BR = 69  # End column (0-indexed)
    COL_AP = 41  # Midpoint Quiz
    COL_BU = 72  # Endpoint Quiz
    COL_START_DATE = 12  # Start Date column (0-indexed)

    EXPECTED_VIDEO_CHAPTERS = 35

    # Completion thresholds
    COMPLETION_25_VIDEOS = 9
    COMPLETION_50_VIDEOS = 18
    COMPLETION_75_VIDEOS = 26
    COMPLETION_100_VIDEOS = 35
    COMPLETION_100_QUIZZES = 2

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.filename = os.path.basename(filepath)
        self.headers = []
        self.rows = []
        self.validation_results = {}
        self.video_chapter_indices = []
        self.quiz_indices = [self.COL_AP, self.COL_BU]

    def excel_col_name(self, n: int) -> str:
        """Convert 0-based index to Excel column name"""
        result = ""
        while n >= 0:
            result = chr(n % 26 + 65) + result
            n = n // 26 - 1
        return result

    def has_24char_id(self, header: str) -> Tuple[bool, str]:
        """Check if header ends with ' - ' followed by exactly 24 alphanumeric characters"""
        match = re.search(r' - ([a-zA-Z0-9]{24})$', header)
        return (True, match.group(1)) if match else (False, None)

    def extract_language(self) -> str:
        """Extract language from filename (between 'AI Samarth - ' and '-timestamp')"""
        # Pattern: AI Samarth - Language-timestamp.csv
        match = re.search(r'AI Samarth - ([^-]+)-\d+\.csv', self.filename, re.IGNORECASE)
        if match:
            return match.group(1).strip()
        return "Unknown"

    def read_csv(self) -> bool:
        """Read entire CSV file"""
        try:
            with open(self.filepath, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                self.headers = next(reader)
                self.rows = list(reader)
            return True
        except Exception as e:
            self.validation_results['error'] = f"Failed to read file: {str(e)}"
            return False

    def identify_video_chapters(self):
        """Identify video chapter column indices in range R to BR (24-char IDs, not quizzes)"""
        self.video_chapter_indices = []

        for i in range(self.COL_R, min(self.COL_BR + 1, len(self.headers))):
            has_24, id_value = self.has_24char_id(self.headers[i])

            if has_24 and 'quiz' not in self.headers[i].lower():
                self.video_chapter_indices.append(i)

    def is_completed(self, cell_value: str) -> bool:
        """Check if a cell indicates completion"""
        if not cell_value:
            return False
        return 'completed' in cell_value.lower()

    def count_completions(self, row: List[str]) -> Tuple[int, int]:
        """Count completed videos and quizzes for a user"""
        videos_completed = 0
        quizzes_completed = 0

        # Count videos
        for idx in self.video_chapter_indices:
            if idx < len(row) and self.is_completed(row[idx]):
                videos_completed += 1

        # Count quizzes
        for idx in self.quiz_indices:
            if idx < len(row) and self.is_completed(row[idx]):
                quizzes_completed += 1

        return videos_completed, quizzes_completed

    def has_started(self, row: List[str]) -> bool:
        """Check if user has started (at least one completion)"""
        # Check videos
        for idx in self.video_chapter_indices:
            if idx < len(row) and self.is_completed(row[idx]):
                return True

        # Check quizzes
        for idx in self.quiz_indices:
            if idx < len(row) and self.is_completed(row[idx]):
                return True

        return False

    def calculate_progress_percentage(self, videos_completed: int, quizzes_completed: int) -> int:
        """Calculate progress percentage based on completion thresholds"""
        # 100% = all 35 videos + all 2 quizzes
        if videos_completed >= self.COMPLETION_100_VIDEOS and quizzes_completed >= self.COMPLETION_100_QUIZZES:
            return 100
        # 75% = at least 26 videos
        elif videos_completed >= self.COMPLETION_75_VIDEOS:
            return 75
        # 50% = at least 18 videos
        elif videos_completed >= self.COMPLETION_50_VIDEOS:
            return 50
        # 25% = at least 9 videos
        elif videos_completed >= self.COMPLETION_25_VIDEOS:
            return 25
        else:
            return 0

    def validate_all(self) -> bool:
        """Run all validations"""
        if not self.read_csv():
            return False

        # Rule 1: Column R contains Pillar
        if self.COL_R >= len(self.headers) or 'pillar' not in self.headers[self.COL_R].lower():
            print(f"  ✗ Validation failed: Column R must contain 'Pillar'")
            return False

        # Identify video chapters in range R to BR
        self.identify_video_chapters()

        # Rule 2: Must have exactly 35 video chapters in range R to BR
        if len(self.video_chapter_indices) != self.EXPECTED_VIDEO_CHAPTERS:
            print(
                f"  ✗ Validation failed: Expected {self.EXPECTED_VIDEO_CHAPTERS} video chapters in range R to BR, found {len(self.video_chapter_indices)}")
            return False

        # Rule 3: Check AP quiz exists
        if self.COL_AP >= len(self.headers) or 'quiz' not in self.headers[self.COL_AP].lower():
            print(f"  ✗ Validation failed: Column AP must contain 'quiz'")
            return False

        # Rule 4: Check BU quiz exists
        if self.COL_BU >= len(self.headers) or 'quiz' not in self.headers[self.COL_BU].lower():
            print(f"  ✗ Validation failed: Column BU must contain 'quiz'")
            return False

        return True

    def process_and_add_columns(self, output_path: str, start_date_filter: Optional[datetime.date] = None, end_date_filter: Optional[datetime.date] = None) -> Dict:
        """Process CSV, add completion columns, and save"""
        if not self.validate_all():
            return None

        # Add new headers
        new_headers = self.headers + ['Videos Completed', 'Quizzes Completed', 'Progress %']

        # Process each row and add completion data
        processed_rows = []
        completion_stats = {
            'total_users': 0, # Will be set after filtering
            'started': 0,
            'started_with_completion': 0,  # Users who completed at least 1 video/quiz
            'only_1_video': 0,
            '25_percent': 0,
            '50_percent': 0,
            '75_percent': 0,
            '100_percent': 0
        }

        # Track users with their start dates and completion status
        user_data = []
        filtered_rows_count = 0
        
        # Total enrolled users = all rows in CSV (regardless of date filter)
        total_enrolled_users = len(self.rows)
        
        for row in self.rows:
            # Get Start Date for filtering
            start_date_str = row[self.COL_START_DATE] if self.COL_START_DATE < len(row) else ""
            date_obj = parse_start_date(start_date_str)
            
            # Apply Date Filter if configured (for processing, but total_users remains all enrolled)
            if start_date_filter and end_date_filter:
                if not date_obj:
                    # Skip rows with no start date if filtering is active
                    continue 
                if not (start_date_filter <= date_obj <= end_date_filter):
                    continue # Skip rows outside range
            
            # Count Row as valid (for filtered processing)
            filtered_rows_count += 1
            
            videos_completed, quizzes_completed = self.count_completions(row)
            progress_pct = self.calculate_progress_percentage(videos_completed, quizzes_completed)

            # Add completion columns
            new_row = row + [str(videos_completed), str(quizzes_completed), str(progress_pct)]
            processed_rows.append(new_row)

            # Store user data for month-wise analysis
            user_data.append({
                'videos_completed': videos_completed,
                'date_info': date_obj, # Now storing date object
                'has_started': self.has_started(row)
            })

            # Check if user has started (based on Start Date column)
            # Count users where Start Date has a valid date value (not "Not Started" or empty)
            if date_obj is not None:
                completion_stats['started'] += 1
            
            # Check if user has actually completed at least 1 video or quiz
            if self.has_started(row):
                completion_stats['started_with_completion'] += 1
            
            # Check if completed exactly one video
            if videos_completed == 1:
                completion_stats['only_1_video'] += 1

            # Update stats
            if progress_pct >= 25:
                completion_stats['25_percent'] += 1
            if progress_pct >= 50:
                completion_stats['50_percent'] += 1
            if progress_pct >= 75:
                completion_stats['75_percent'] += 1
            if progress_pct == 100:
                completion_stats['100_percent'] += 1

        # Set total users to all enrolled users (total rows in CSV, regardless of date filter)
        completion_stats['total_users'] = total_enrolled_users

        # Calculate month-wise "At Least 1 Video" metrics dynamically
        # First, detect all unique year-month combinations present in the data
        year_month_pairs = set()
        years_in_data = []
        for user in user_data:
            if user['date_info']:
                year_month_pairs.add((user['date_info'].year, user['date_info'].month))
                years_in_data.append(user['date_info'].year)
        
        if not year_month_pairs:
            # No date data available
            return completion_stats
        
        # Sort year-month pairs chronologically (by year, then by month)
        sorted_year_months = sorted(year_month_pairs)
        
        # Get year range
        if years_in_data:
            min_year = min(years_in_data)
            max_year = max(years_in_data)
        else:
            min_year = max_year = 2025
        
        # Month names for dynamic column creation
        month_names = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 
                      'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
        
        # Initialize month-wise stats dynamically with year-month keys
        for year, month_num in sorted_year_months:
            month_name = month_names[month_num - 1]
            # Include year in column name to handle multiple years
            completion_stats[f'at_least_1_video_cumulative_{year}_{month_name}'] = 0
            completion_stats[f'at_least_1_video_monthly_{year}_{month_name}'] = 0
        
        # Calculate month end dates for cumulative counts (for each year present)
        from datetime import timedelta
        from calendar import monthrange
        
        # Process users for month-wise metrics
        for user in user_data:
            if user['videos_completed'] >= 1 and user['date_info']:
                user_start_date = user['date_info']
                user_month = user_start_date.month
                user_year = user_start_date.year
                month_name = month_names[user_month - 1]
                
                # Cumulative counts: users who started on or before the end of each month
                # AND have completed at least 1 video
                for cum_year, cum_month_num in sorted_year_months:
                    cum_month_name = month_names[cum_month_num - 1]
                    cum_key = f'at_least_1_video_cumulative_{cum_year}_{cum_month_name}'
                    
                    # Get the last day of the month for the year we're checking
                    last_day = monthrange(cum_year, cum_month_num)[1]
                    month_end = datetime(cum_year, cum_month_num, last_day).date()
                    
                    # Count user if they started on or before the end of this month
                    if user_start_date <= month_end:
                        completion_stats[cum_key] += 1
                
                # Monthly counts: users who started in that specific month only
                monthly_key = f'at_least_1_video_monthly_{user_year}_{month_name}'
                if monthly_key in completion_stats:
                    completion_stats[monthly_key] += 1

        # Store user data summary for post-processing (to handle missing month columns)
        # Store list of (start_date, videos_completed) for users with completions
        completion_stats['_user_data_summary'] = [
            (user['date_info'], user['videos_completed']) 
            for user in user_data 
            if user['videos_completed'] >= 1 and user['date_info']
        ]

        # Write to new CSV
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(new_headers)
            writer.writerows(processed_rows)

        return completion_stats

    def print_summary(self, stats: Dict):
        """Print completion summary"""
        print(f"\n{'=' * 80}")
        print(f"File: {self.filename}")
        print(f"{'=' * 80}")
        print(f"Total Users: {stats['total_users']}")
        print(f"Started: {stats['started']} users ({stats['started'] / stats['total_users'] * 100:.1f}%)")
        print(f"Started (Completed ≥1 Video/Quiz): {stats['started_with_completion']} users ({stats['started_with_completion'] / stats['total_users'] * 100:.1f}%)")
        print(f"Only 1 Video: {stats['only_1_video']} users ({stats['only_1_video'] / stats['total_users'] * 100:.1f}%)")
        print(f"25% Completion: {stats['25_percent']} users ({stats['25_percent'] / stats['total_users'] * 100:.1f}%)")
        print(f"50% Completion: {stats['50_percent']} users ({stats['50_percent'] / stats['total_users'] * 100:.1f}%)")
        print(f"75% Completion: {stats['75_percent']} users ({stats['75_percent'] / stats['total_users'] * 100:.1f}%)")
        print(
            f"100% Completion: {stats['100_percent']} users ({stats['100_percent'] / stats['total_users'] * 100:.1f}%)")


def normalize_month_columns(all_stats: List[Dict]):
    """
    Ensure all stats dictionaries have columns for all year-months found across all files.
    This fixes the issue where a file doesn't have a column for a month that exists in other files.
    Recalculates cumulative values using stored user data.
    """
    import re
    from calendar import monthrange
    
    month_names = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 
                  'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
    
    # Collect all unique year-month pairs from all stats
    all_year_months = set()
    for stats in all_stats:
        for col in stats.keys():
            # Check cumulative columns
            match = re.match(r'at_least_1_video_cumulative_(\d+)_(\w+)', col)
            if match:
                year = int(match.group(1))
                month_abbr = match.group(2)
                if month_abbr in month_names:
                    all_year_months.add((year, month_names.index(month_abbr) + 1))
    
    if not all_year_months:
        return
    
    # For each stats dict, add missing columns and recalculate cumulative values
    for stats in all_stats:
        # Get user data summary if available
        user_data_summary = stats.get('_user_data_summary', [])
        
        # For each year-month (including missing ones), calculate cumulative value
        for year, month_num in sorted(all_year_months):
            month_name = month_names[month_num - 1]
            cum_key = f'at_least_1_video_cumulative_{year}_{month_name}'
            mon_key = f'at_least_1_video_monthly_{year}_{month_name}'
            
            # Initialize if missing
            if cum_key not in stats:
                stats[cum_key] = 0
            if mon_key not in stats:
                stats[mon_key] = 0
            
            # Recalculate cumulative using user data
            if user_data_summary:
                # Get the last day of the month
                last_day = monthrange(year, month_num)[1]
                month_end = datetime(year, month_num, last_day).date()
                
                # Count users who started on or before the end of this month
                cumulative_count = sum(
                    1 for start_date, videos_completed in user_data_summary
                    if start_date and start_date <= month_end
                )
                stats[cum_key] = cumulative_count
                
                # Calculate monthly count (users who started in this specific month)
                monthly_count = sum(
                    1 for start_date, videos_completed in user_data_summary
                    if start_date and start_date.year == year and start_date.month == month_num
                )
                stats[mon_key] = monthly_count
        
        # Remove the temporary user data summary (cleanup)
        if '_user_data_summary' in stats:
            del stats['_user_data_summary']


def create_summary_excel(all_stats: List[Dict], output_path: str):
    """Create summary Excel with completion statistics"""

    # Sort by total_users in descending order (highest first)
    all_stats_sorted = sorted(all_stats, key=lambda x: x['total_users'], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Completion Summary"

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers - Changed "File Name" to "Course Language"
    headers = ['Course Language', 'Total Users', 'Started', 'Started (Completed ≥1 Video/Quiz)', 'Only 1 Video', '25% Completion', '50% Completion', '75% Completion',
               '100% Completion']
    ws.append(headers)

    # Style header row - left aligned
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border

    # Add data rows - use sorted data
    for stats in all_stats_sorted:
        ws.append([
            stats['language'],
            stats['total_users'],
            stats['started'],
            stats.get('started_with_completion', 0),
            stats['only_1_video'],
            stats['25_percent'],
            stats['50_percent'],
            stats['75_percent'],
            stats['100_percent']
        ])

    # Calculate totals
    total_row = ['OVERALL TOTALS']
    for col_idx in range(2, 10):  # Columns B to I (Total Users to 100% Completion)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        start_row = 2
        end_row = len(all_stats_sorted) + 1
        total_row.append(f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")

    ws.append(total_row)
    total_row_idx = len(all_stats_sorted) + 2

    # Style total row - left aligned
    for cell in ws[total_row_idx]:
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border

    # Style data rows - all left aligned
    for row_idx in range(2, total_row_idx):
        for cell in ws[row_idx]:
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['D'].width = 35  # Wider for "Started (Completed ≥1 Video/Quiz)"
    for col in ['B', 'C', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 18

    # Add month-wise analysis sheets
    _add_monthwise_sheets(wb, all_stats_sorted)
    
    # Save workbook
    wb.save(output_path)
    print(f"\n✓ Summary Excel created: {output_path}")


def _add_monthwise_sheets(wb, all_stats_sorted):
    """Add month-wise 'At Least 1 Video' analysis sheets to workbook"""
    
    # Define styles (same as main sheet)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Month names for dynamic column detection
    import re
    month_names = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 
                  'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
    month_display_names = ['January', 'February', 'March', 'April', 'May', 'June',
                          'July', 'August', 'September', 'October', 'November', 'December']
    
    # Detect all cumulative month columns present in the data (format: at_least_1_video_cumulative_YYYY_month)
    cum_month_cols = []
    cum_display_info = []  # List of (col_name, year, month_num, display_name)
    if all_stats_sorted:
        for col in all_stats_sorted[0].keys():
            match = re.match(r'at_least_1_video_cumulative_(\d+)_(\w+)', col)
            if match:
                year = int(match.group(1))
                month_abbr = match.group(2)
                if month_abbr in month_names:
                    month_idx = month_names.index(month_abbr)
                    month_num = month_idx + 1
                    display_name = f'Up to {month_display_names[month_idx]} {year} End'
                    cum_month_cols.append(col)
                    cum_display_info.append((col, year, month_num, display_name))
        
        # Sort by year, then by month
        cum_display_info.sort(key=lambda x: (x[1], x[2]))
        cum_month_cols = [info[0] for info in cum_display_info]
        cum_display_names = [info[3] for info in cum_display_info]
    
    # Detect all monthly month columns present in the data (format: at_least_1_video_monthly_YYYY_month)
    mon_month_cols = []
    mon_display_info = []  # List of (col_name, year, month_num, display_name)
    if all_stats_sorted:
        for col in all_stats_sorted[0].keys():
            match = re.match(r'at_least_1_video_monthly_(\d+)_(\w+)', col)
            if match:
                year = int(match.group(1))
                month_abbr = match.group(2)
                if month_abbr in month_names:
                    month_idx = month_names.index(month_abbr)
                    month_num = month_idx + 1
                    display_name = f'{month_display_names[month_idx]} {year} Only'
                    mon_month_cols.append(col)
                    mon_display_info.append((col, year, month_num, display_name))
        
        # Sort by year, then by month
        mon_display_info.sort(key=lambda x: (x[1], x[2]))
        mon_month_cols = [info[0] for info in mon_display_info]
        mon_display_names = [info[3] for info in mon_display_info]
    
    # Sheet 1: Cumulative "At Least 1 Video" (Start to Month End)
    if cum_month_cols:
        ws_cumulative = wb.create_sheet("At Least 1 Video - Cumulative")
        
        headers_cumulative = ['Course Language'] + cum_display_names
        ws_cumulative.append(headers_cumulative)
        
        # Style header row
        for cell in ws_cumulative[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
        
        # Add data rows
        for stats in all_stats_sorted:
            row_data = [stats['language']]
            for col_name in cum_month_cols:
                row_data.append(stats.get(col_name, 0))
            ws_cumulative.append(row_data)
        
        # Add totals row
        total_row_cumulative = ['OVERALL TOTALS']
        for col_idx in range(2, len(cum_month_cols) + 2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            start_row = 2
            end_row = len(all_stats_sorted) + 1
            total_row_cumulative.append(f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        
        ws_cumulative.append(total_row_cumulative)
        total_row_idx_cumulative = len(all_stats_sorted) + 2
        
        # Style total row
        for cell in ws_cumulative[total_row_idx_cumulative]:
            cell.fill = total_fill
            cell.font = total_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
        
        # Style data rows
        for row_idx in range(2, total_row_idx_cumulative):
            for cell in ws_cumulative[row_idx]:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Adjust column widths
        ws_cumulative.column_dimensions['A'].width = 20
        for col_idx in range(2, len(cum_month_cols) + 2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws_cumulative.column_dimensions[col_letter].width = 18
    
    # Sheet 2: Monthly "At Least 1 Video" (Month Only)
    if mon_month_cols:
        ws_monthly = wb.create_sheet("At Least 1 Video - Monthly")
        
        headers_monthly = ['Course Language'] + mon_display_names
        ws_monthly.append(headers_monthly)
        
        # Style header row
        for cell in ws_monthly[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
        
        # Add data rows
        for stats in all_stats_sorted:
            row_data = [stats['language']]
            for col_name in mon_month_cols:
                row_data.append(stats.get(col_name, 0))
            ws_monthly.append(row_data)
        
        # Add totals row
        total_row_monthly = ['OVERALL TOTALS']
        for col_idx in range(2, len(mon_month_cols) + 2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            start_row = 2
            end_row = len(all_stats_sorted) + 1
            total_row_monthly.append(f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
        
        ws_monthly.append(total_row_monthly)
        total_row_idx_monthly = len(all_stats_sorted) + 2
        
        # Style total row
        for cell in ws_monthly[total_row_idx_monthly]:
            cell.fill = total_fill
            cell.font = total_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
        
        # Style data rows
        for row_idx in range(2, total_row_idx_monthly):
            for cell in ws_monthly[row_idx]:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Adjust column widths
        ws_monthly.column_dimensions['A'].width = 20
        for col_idx in range(2, len(mon_month_cols) + 2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws_monthly.column_dimensions[col_letter].width = 18


def find_aisamarth_files(source_path: Path) -> List[str]:
    """Find all files containing 'AI Samarth' (case insensitive) in specified folder"""
    files = []

    if not source_path.exists():
        print(f"❌ Folder not found: {source_path}")
        return files

    for file in source_path.iterdir():
        if file.is_file() and 'ai samarth' in file.name.lower():
            files.append(str(file))

    return files


def extract_language_from_filename(filename: str) -> str:
    """Extract language from filename"""
    match = re.search(r'AI Samarth - ([^-]+)-\d+\.csv', filename, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return None


def validate_language_files(files: List[str]) -> Tuple[bool, str]:
    """Validate that files contain exactly the required 5 languages with no duplicates"""

    REQUIRED_LANGUAGES = {'English', 'Hindi', 'Marathi', 'Bengali', 'Odia'}

    # Extract languages from filenames
    found_languages = []
    language_to_file = {}

    for file in files:
        filename = os.path.basename(file)
        language = extract_language_from_filename(filename)

        if language:
            found_languages.append(language)
            if language in language_to_file:
                language_to_file[language].append(filename)
            else:
                language_to_file[language] = [filename]

    # Check for duplicates
    duplicates = {lang: files for lang, files in language_to_file.items() if len(files) > 1}
    if duplicates:
        error_msg = "\n❌ HARD RULE VIOLATION: Duplicate language files found:\n"
        for lang, dup_files in duplicates.items():
            error_msg += f"   {lang}: {len(dup_files)} files\n"
            for f in dup_files:
                error_msg += f"      - {f}\n"
        return False, error_msg

    # Convert to set for comparison (case-insensitive)
    found_languages_set = {lang for lang in found_languages}

    # Check if we have exactly the required languages
    missing_languages = REQUIRED_LANGUAGES - found_languages_set
    extra_languages = found_languages_set - REQUIRED_LANGUAGES

    if missing_languages or extra_languages:
        error_msg = "\n❌ HARD RULE VIOLATION: Invalid language files:\n"
        error_msg += f"   Required languages: {', '.join(sorted(REQUIRED_LANGUAGES))}\n"
        error_msg += f"   Found languages: {', '.join(sorted(found_languages_set)) if found_languages_set else 'None'}\n"

        if missing_languages:
            error_msg += f"   Missing: {', '.join(sorted(missing_languages))}\n"
        if extra_languages:
            error_msg += f"   Extra/Invalid: {', '.join(sorted(extra_languages))}\n"

        return False, error_msg

    return True, "✓ All required language files present with no duplicates"


def main():
    """Main function"""
    # Windows path: C:\Users\<YourUsername>\Downloads\PyCharm - AI Samarth
    # Path.home() automatically detects the user's home directory on Windows
    project_root = Path(__file__).parent
    source_path = project_root / "data_files"

    print("=" * 100)
    print("AI SAMARTH CSV PROCESSOR WITH COMPLETION TRACKING")
    print("=" * 100)
    print(f"\nSearching for AI Samarth files in: {source_path}")

    # Find files
    files = find_aisamarth_files(source_path)

    if not files:
        print("\n❌ No files containing 'AI Samarth' found in the specified folder")
        return

    # HARD RULE: Must have exactly 5 files
    if len(files) != 5:
        print(f"\n❌ HARD RULE VIOLATION: Folder must contain exactly 5 AI Samarth files")
        print(f"   Found: {len(files)} file(s)")
        print(f"   Required: 5 files")
        print("\n🚫 Processing ignored. Exiting...")
        return

    print(f"\n✓ Found {len(files)} file(s) containing 'AI Samarth':\n")
    for i, file in enumerate(files, 1):
        print(f"  {i}. {os.path.basename(file)}")

    # HARD RULE: Validate language files
    is_valid, validation_msg = validate_language_files(files)
    if not is_valid:
        print(validation_msg)
        print("\n🚫 Processing ignored. Exiting...")
        return

    print(f"\n{validation_msg}")

    # Create output folders
    output_folder = project_root / "output"
    csv_output_folder = output_folder / "Processed_CSVs"
    output_folder.mkdir(exist_ok=True)
    csv_output_folder.mkdir(exist_ok=True)

    print(f"\n✓ Output folder: {output_folder}")
    print(f"✓ Processed CSVs will be in: {csv_output_folder}")

    # Process each file
    print("\nPROCESSING FILES")
    print("=" * 100)

    all_stats = []

    for file in files:
        processor = AISmarthProcessor(file)

        # Generate output filename
        base_name = os.path.splitext(os.path.basename(file))[0]
        output_csv = csv_output_folder / f"{base_name}_processed.csv"

        print(f"\nProcessing: {os.path.basename(file)}")

        stats = processor.process_and_add_columns(str(output_csv))

        if stats:
            print(f"✓ Saved: {output_csv.name}")
            processor.print_summary(stats)

            # Extract language and add to stats
            stats['language'] = processor.extract_language()
            stats['filename'] = os.path.basename(file)
            all_stats.append(stats)
        else:
            print(f"✗ Failed to process (validation failed)")

    # Create summary Excel
    if all_stats:
        # Normalize month columns across all files
        normalize_month_columns(all_stats)
        
        print("\n" + "=" * 100)
        print("CREATING SUMMARY EXCEL")
        print("=" * 100)

        summary_excel = output_folder / "AI_Samarth_Summary.xlsx"
        create_summary_excel(all_stats, str(summary_excel))

        # Print overall summary
        print("\n" + "=" * 100)
        print("OVERALL SUMMARY")
        print("=" * 100)

        total_users = sum(s['total_users'] for s in all_stats)
        total_started = sum(s['started'] for s in all_stats)
        total_started_with_completion = sum(s.get('started_with_completion', 0) for s in all_stats)
        total_only_1 = sum(s['only_1_video'] for s in all_stats)
        total_25 = sum(s['25_percent'] for s in all_stats)
        total_50 = sum(s['50_percent'] for s in all_stats)
        total_75 = sum(s['75_percent'] for s in all_stats)
        total_100 = sum(s['100_percent'] for s in all_stats)

        print(f"\nTotal Users Across All Files: {total_users}")
        print(f"Started: {total_started} users ({total_started / total_users * 100:.1f}%)")
        print(f"Started (Completed ≥1 Video/Quiz): {total_started_with_completion} users ({total_started_with_completion / total_users * 100:.1f}%)")
        print(f"Only 1 Video: {total_only_1} users ({total_only_1 / total_users * 100:.1f}%)")
        print(f"25% Completion: {total_25} users ({total_25 / total_users * 100:.1f}%)")
        print(f"50% Completion: {total_50} users ({total_50 / total_users * 100:.1f}%)")
        print(f"75% Completion: {total_75} users ({total_75 / total_users * 100:.1f}%)")
        print(f"100% Completion: {total_100} users ({total_100 / total_users * 100:.1f}%)")

        print("\n" + "=" * 100)
        print("✓✓✓ PROCESSING COMPLETE ✓✓✓")
        print("=" * 100)
        print(f"\nProcessed files location: {output_folder}")
        print(f"Summary Excel: {summary_excel}")


if __name__ == "__main__":
    main()